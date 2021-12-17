# Created by: Alice Castillo

# Description:
from enum import Enum
from datetime import datetime
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill


class EvalColor(Enum):
    Title = "000000"
    Lower = "0066CC"
    Equal = "00CC00"
    Greater = "FF3333"
    Error = "808080"



class LogFile():
    def __init__(self, filename, cur_time):
        self.filename = f'plant_specs/subject_logs/subject1/{filename}.xlsx'
        self.cur_time = cur_time
        self.createExcelFile()
        # self.wb = self.createExcelFile()


    def createExcelFile(self) -> None:
        # create file
        try:
            wb = load_workbook(self.filename)
        except FileNotFoundError as e:
            # Create Excel file through xlsxwriter
            workbook = xlsxwriter.Workbook(self.filename)
            workbook.close()

            # Now modify to add sheets for all sensors (This could be generalized in future)
            wb = load_workbook(self.filename)
            sheets = ["Humidity", "Temperature", "UV Level"]
            for sheet in sheets:
                ws = wb.create_sheet(title=sheet)
                header = ws.row_dimensions[1]
                header.font = Font(bold=True, color="FFFFFF")
                header.fill = PatternFill("solid", fgColor="000000")

                x_axis = ["Time", "Average", "Median"]
                if sheet == "UV Level":
                    x_axis.append("Forecasted")

                for count in range(len(x_axis)):
                    ws["A{}".format(count+1)] = x_axis[count]
                    cell = ws["A{}".format(count+1)]
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=EvalColor.Title.value)
            wb.save(self.filename)
            wb = load_workbook(self.filename)
            wb.close()

            # Remove default-add "Sheet1"
            wb = load_workbook(self.filename)
            if wb.sheetnames[0] == "Sheet1":
                wb.remove(wb[wb.sheetnames[0]])  # Remove "Sheet1"
            wb.save(self.filename)

        # Close and exit
        # return wb
        wb.close()
        # return wb

    def colNumToLetters(self, col_num: int) -> str:
        """ Converts a column num to Excel's lettering system.
            Ex: Col 1 = 'A'
        """
        if col_num > 26:
            return "{}{}".format(chr(col_num // 26 + 64), chr(col_num % 26 + 64))
        return "{}".format(chr(col_num % 26 + 64))

    def applyHeaderStyling(self):
        wb = load_workbook(self.filename)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws["1:1"]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="000000")
        wb.save(self.filename)
        wb.close()


    def insertCol(self, sheet, data, colors):
        wb = load_workbook(self.filename)
        ws = wb[sheet]
        next_col = wb[sheet].max_column + 1
        ws.insert_cols(next_col)

        # add title
        ws["{}1".format(self.colNumToLetters(next_col))] = datetime.fromtimestamp(self.cur_time).strftime("%H:%M")

        for count, d in enumerate(data):
            ws["{}{}".format(self.colNumToLetters(next_col), count + 2)] = d
            cell = ws["{}{}".format(self.colNumToLetters(next_col), count + 2)]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=colors[count].value)
        wb.save(self.filename)
        wb.close()
